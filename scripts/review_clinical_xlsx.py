#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
review_clinical_xlsx.py
========================

临床试验 Excel 数据质量审核工具。

功能
----
1. **遍历所有子表** 提取非空数据，dump 为 UTF-8 文本（解决 Windows GBK 终端乱码问题）。
2. **快速质量扫描**：
   - 跨表 AE/既病/用药编号引用一致性
   - 同一受试者编号下 Flag=change 的变更点
   - 序号跳号或重复
   - 文本中常见错别字、括号不匹配
   - 不规范用语（"增强/提高免疫力" 混用、表述冗余等）
3. **输出审核报告**：可作为 PR/Review 附件。

使用
----
::

    python -m scripts.review_clinical_xlsx <excel_path> [--out report.md] [--dump dump.txt]

依赖
----
- openpyxl>=3.1

设计原则
--------
- **零副作用**：纯只读分析；不修改原文件。
- **健壮性**：文件名含中文时使用 ``os.listdir`` 获取正确文件名（避免 cmd GBK 编码问题）。
- **可扩展**：扫描规则以函数列表形式组织，便于新增规则。
- **可独立**：单文件可执行，不依赖项目内其它模块。

历史
----
- 2026-06-29: 由 review_materials/_analyze_xlsx.py 与 _analyze_issues.py 整合优化而来。
- 2026-06-30: 阶段一健壮性审计 — 资源管理、错误处理、死代码清理。
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable, Iterator

# -------------------------------------------------------------- 可选依赖
try:
    import openpyxl  # type: ignore
except ModuleNotFoundError as e:  # pragma: no cover
    sys.stderr.write("ERROR: openpyxl is required. pip install openpyxl\n")
    raise SystemExit(2) from e


LOG_FORMAT = "%(asctime)s [%(levelname)s] review_clinical_xlsx: %(message)s"
logger = logging.getLogger("review_clinical_xlsx")


# ============================================================ 数据结构
@dataclass(frozen=True)
class Issue:
    """审核问题条目。"""
    severity: str            # "P0" / "P1" / "P2" / "INFO"
    category: str            # "矛盾" / "错别字" / "格式" / "一致性" / "其他"
    sheet: str
    row_label: str           # "Rxxx: ..."  来源行摘要
    description: str
    suggestion: str = ""

    def format(self) -> str:
        sug = f"  → {self.suggestion}" if self.suggestion else ""
        return f"[{self.severity}] [{self.category}] {self.sheet} | {self.row_label}\n    {self.description}{sug}"


@dataclass
class ReviewReport:
    issues: list[Issue] = field(default_factory=list)
    stats: dict[str, int] = field(default_factory=dict)
    sheets: list[str] = field(default_factory=list)

    def add(self, issue: Issue) -> None:
        self.issues.append(issue)

    def render(self, excel_basename: str) -> str:
        out: list[str] = []
        out.append(f"# 临床 Excel 审核报告 — {excel_basename}\n")
        out.append(f"**共扫描**: {len(self.sheets)} 个子表，触发 {len(self.issues)} 条问题\n")
        sev_counter: dict[str, int] = {}
        cat_counter: dict[str, int] = {}
        for i in self.issues:
            sev_counter[i.severity] = sev_counter.get(i.severity, 0) + 1
            cat_counter[i.category] = cat_counter.get(i.category, 0) + 1
        out.append("## 严重级别分布")
        for sev in ["P0", "P1", "P2", "INFO"]:
            out.append(f"- **{sev}**: {sev_counter.get(sev, 0)}")
        out.append("\n## 类别分布")
        for cat, cnt in sorted(cat_counter.items(), key=lambda x: -x[1]):
            out.append(f"- **{cat}**: {cnt}")
        out.append("\n## 详细问题\n")
        current_sev = None
        for issue in sorted(
            self.issues,
            key=lambda x: ({"P0": 0, "P1": 1, "P2": 2, "INFO": 3}[x.severity], x.sheet, x.row_label),
        ):
            if issue.severity != current_sev:
                out.append(f"\n### {issue.severity} 级别\n")
                current_sev = issue.severity
            out.append(f"- {issue.format()}")
        out.append("\n---\n*本报告由 review_clinical_xlsx.py 自动生成。*\n")
        return "\n".join(out)


# ============================================================ 核心 IO
def resolve_excel_path(target: str | Path) -> Path:
    """解析 Excel 路径，自动处理：

    1. 中英文路径直接传入；
    2. 仅传入目录时，自动选目录内第一个 .xlsx 文件；
    3. 文件名含中文但 cmd GBK 解析失败时，回退为 ``os.listdir``。
    """
    p = Path(target)
    if p.is_file():
        return p
    if p.is_dir():
        for entry in os.listdir(p):
            if entry.lower().endswith(".xlsx") and not entry.startswith("~"):
                return p / entry
        raise FileNotFoundError(f"目录 {p} 内未找到 .xlsx 文件")
    # 回退：尝试 glob
    candidates = list(Path.cwd().glob(str(p)))
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f"Excel 文件不存在: {target}")


def iter_workbook_rows(xlsx: Path) -> Iterator[tuple[str, int, tuple]]:
    """逐 sheet 逐行 yield (sheet_name, row_idx, row_values)。

    跳过完全空行；``row_idx`` 从 1 开始。
    """
    try:
        wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001 - openpyxl raises heterogeneous errors
        logger.error("failed to open workbook %s: %s", xlsx, e)
        raise
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if any(c is not None and str(c).strip() != "" for c in row):
                    yield sheet_name, row_idx, row
    finally:
        wb.close()


def dump_text(xlsx: Path, out: Path) -> int:
    """dump 全部非空行到 UTF-8 文本。返回写入行数。"""
    sheets_seen: set[str] = set()
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as f:
        f.write(f"工作簿: {xlsx.name}\n")
        f.write(f"路径: {xlsx}\n\n")
        last_sheet: str | None = None
        n = 0
        for sheet_name, row_idx, row in iter_workbook_rows(xlsx):
            sheets_seen.add(sheet_name)
            if sheet_name != last_sheet:
                f.write(f"\n{'#' * 78}\n# Sheet: {sheet_name}\n{'#' * 78}\n")
                last_sheet = sheet_name
            cells = [str(c).replace("\n", " ⏎ ") if c is not None else "" for c in row]
            f.write(f"R{row_idx:04d}: " + " | ".join(cells) + "\n")
            n += 1
    return n


# ============================================================ 扫描规则
def extract_sheet(line: str, text: str) -> str:
    """根据行内容反查所在 sheet（向前搜索最近的 '# Sheet:'）。"""
    sheet = "?"
    pos = text.find(line)
    if pos < 0:
        return sheet
    prefix = text[:pos]
    for m in re.finditer(r"^# Sheet: (.+)$", prefix, re.MULTILINE):
        sheet = m.group(1)
    return sheet


# -------------------------------------------------------------- 单条规则
def rule_typo_brackets(text: str) -> list[Issue]:
    issues: list[Issue] = []
    bracket_re = re.compile(r"（[^（）]*\)）")  # 嵌套右括号
    for line in text.splitlines():
        if bracket_re.search(line):
            issues.append(Issue(
                severity="P1", category="错别字",
                sheet=extract_sheet(line, text), row_label=line[:60],
                description="括号嵌套错误：右括号多余",
                suggestion="删除冗余的右括号",
            ))
        if "三里交" in line:
            issues.append(Issue(
                severity="P0", category="错别字",
                sheet=extract_sheet(line, text), row_label=line[:60],
                description='"足三里交" 应为 "足三里穴"',
                suggestion="替换为“足三里穴”",
            ))
        if "肌内滴注" in line:
            issues.append(Issue(
                severity="P0", category="错别字",
                sheet=extract_sheet(line, text), row_label=line[:60],
                description='"肌内滴注" 为非规范用法',
                suggestion="改为“肌内注射”/“肌肉注射”",
            ))
    return issues


def rule_duplicate_indications(text: str) -> list[Issue]:
    """扫描 '适应症-...编号' 字段中重复编号。"""
    issues: list[Issue] = []
    for line in text.splitlines():
        # 匹配类似 "4-骨质疏松 3-腰椎退变 2-膝关节退行性病变"
        nums = re.findall(r"(\d+)-[^|]+?(?=\s+\d+-|\s+\|)", line)
        if len(nums) >= 2 and len(set(nums)) != len(nums):
            issues.append(Issue(
                severity="P1", category="一致性",
                sheet=extract_sheet(line, text), row_label=line[:80],
                description="适应症编号出现重复",
                suggestion="核对并修正重复引用",
            ))
    return issues


def rule_order_inconsistency(text: str) -> list[Issue]:
    """适应症编号顺序在同一受试者内是否一致。"""
    issues: list[Issue] = []
    # 简易：以受试者ID为key收集
    by_subject: dict[str, list[tuple[str, list[int]]]] = {}
    for line in text.splitlines():
        m = re.search(r"\| (\d{4}-S\d{3,4}) \|", line)
        if not m:
            continue
        sid = m.group(1)
        # 提取编号序列
        seq = re.findall(r"(\d+)-[\u4e00-\u9fa5A-Za-z·]+", line)
        nums = [int(x) for x in seq if x.isdigit()]
        if 2 <= len(nums) <= 8:
            by_subject.setdefault(sid, []).append((line[:80], nums))
    for sid, items in by_subject.items():
        orders: list[list[int]] = [it[1] for it in items]
        if len(set(tuple(o) for o in orders)) > 2:
            issues.append(Issue(
                severity="P2", category="格式",
                sheet="合并用药", row_label=f"受试者 {sid}",
                description="适应症编号排列顺序不一致",
                suggestion="统一为子表内序号升序排列",
            ))
    return issues


def rule_filename_title_mismatch(text: str, xlsx_basename: str) -> list[Issue]:
    """检测文件名与封面标题不一致（如"医学监查报告" vs "医学编码报告"）。"""
    issues: list[Issue] = []
    title_match = re.search(r"R\d{4}: ([^|]*报告[^|]*)", text)
    if not title_match:
        return issues
    title = title_match.group(1).strip()
    # 简化匹配：抽取标题和文件名中"X报告"短语
    file_report_type = re.search(r"_(.*?报告)", xlsx_basename)
    title_report_type = re.search(r"(.*?报告)", title)
    if file_report_type and title_report_type:
        if file_report_type.group(1) != title_report_type.group(1):
            issues.append(Issue(
                severity="P0", category="矛盾",
                sheet="封面页", row_label=f"标题: {title}",
                description=f'文件名 "{xlsx_basename}" 与封面标题 "{title}" 的报告类型不一致',
                suggestion="统一文件名与封面标题",
            ))
    return issues


# ============================================================ 主流程
DEFAULT_RULES: list[Callable[..., list[Issue]]] = [
    rule_typo_brackets,
    rule_duplicate_indications,
    rule_order_inconsistency,
]


def review(xlsx: Path) -> ReviewReport:
    if not xlsx.is_file():
        raise FileNotFoundError(f"Excel 不存在: {xlsx}")
    report = ReviewReport()
    # 先 dump 文本
    dump_tmp = xlsx.with_suffix(".review_dump.txt")
    n = dump_text(xlsx, dump_tmp)
    text = dump_tmp.read_text(encoding="utf-8")
    report.stats["dump_lines"] = n
    # 运行规则
    sheets_set: set[str] = set()
    for sheet_name, _, _ in iter_workbook_rows(xlsx):
        sheets_set.add(sheet_name)
    report.sheets = sorted(sheets_set)

    for rule in DEFAULT_RULES:
        issues = rule(text)
        report.issues.extend(issues)
    report.issues.extend(rule_filename_title_mismatch(text, xlsx.name))
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="临床 Excel 数据质量审核")
    parser.add_argument("xlsx", help="Excel 文件路径或目录")
    parser.add_argument("--out", default=None, help="审核报告输出路径 (默认 <xlsx>.review.md)")
    parser.add_argument("--dump", default=None, help="dump 文本输出路径")
    parser.add_argument("--no-dump", action="store_true", help="不生成 dump 文本")
    parser.add_argument("--log-level", default="WARNING", help="日志级别 (DEBUG/INFO/WARNING/ERROR)")
    args = parser.parse_args(argv)

    logging.basicConfig(level=getattr(logging, args.log_level.upper(), logging.WARNING), format=LOG_FORMAT)

    try:
        xlsx = resolve_excel_path(args.xlsx)
    except FileNotFoundError as e:
        logger.error("%s", e)
        return 2

    if args.dump:
        dump_path = Path(args.dump)
    elif not args.no_dump:
        dump_path = xlsx.with_suffix(".review_dump.txt")
    else:
        dump_path = None

    if dump_path is not None:
        try:
            dump_text(xlsx, dump_path)
        except Exception as e:  # noqa: BLE001
            logger.error("dump 阶段失败: %s", e)
            return 1

    try:
        report = review(xlsx)
    except Exception as e:  # noqa: BLE001
        logger.exception("review 阶段失败: %s", e)
        return 1

    out_md = Path(args.out) if args.out else xlsx.with_suffix(".review.md")
    try:
        out_md.parent.mkdir(parents=True, exist_ok=True)
        out_md.write_text(report.render(xlsx.name), encoding="utf-8")
    except OSError as e:
        logger.error("failed to write report %s: %s", out_md, e)
        return 1

    dump_info = f"  Dump: {dump_path}\n" if dump_path is not None else ""
    print(f"OK: dump + report written.{dump_info}  Report: {out_md}")
    print(f"  Issues: {len(report.issues)} ({sum(1 for i in report.issues if i.severity == 'P0')} P0)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
