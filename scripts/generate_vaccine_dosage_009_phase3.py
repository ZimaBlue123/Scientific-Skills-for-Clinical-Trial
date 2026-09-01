#!/usr/bin/env python3
"""
generate_vaccine_dosage_009_phase3.py
=====================================

计算 009 / 009B 两个Ⅲ期乙肝疫苗项目的**试验疫苗**用量，输出 Excel 汇总表。

数据来源
--------
- review_materials/009 产品用量/远大赛威信重组乙型肝炎疫苗（汉逊酵母，CpG和铝佐剂）Ⅲ期-临床方案-V1.0-20260526-clean.docx
- review_materials/009 产品用量/远大赛威信重组乙型肝炎疫苗（铝佐剂）Ⅲ期临床方案（18岁及以上）-V1.0-20210922-解密.docx

口径
----
- 仅统计试验疫苗（对照疫苗除外）。
- 留样用量为假设值：按《药品生产质量管理规范》临床试验用药品附录（2022）第三十六条
  “每批留样 ≥ 2 次全检量 + 至少 1 件最小包装”，假设重组乙肝疫苗一次全检量约 9~10 支、
  按 1 个生产批次计，留样取整 20 支/项目（可调整）。

计算
----
总体用量 = 基础用量 + 备用用量 + 留样用量（1 支 = 1 剂，0.5 ml/支单剂量）
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = (
    r"E:/Cursor Project/2-Scientific-Skills-for-Clinical_Trial/"
    r"review_materials/009 产品用量/009系列Ⅲ期试验疫苗用量汇总.xlsx"
)

# -------------------------------------------------------------- 样式常量
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
BODY_FONT = Font(name="微软雅黑", size=10)
PROJ_FONT = Font(name="微软雅黑", size=10, bold=True)
NOTE_FONT = Font(name="微软雅黑", size=9, color="595959")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "疫苗用量汇总"

    headers = [
        "项目",
        "样本量（例）",
        "免疫程序",
        "备用疫苗比例",
        "留样用量（支）",
        "基础用量（剂）",
        "备用用量（剂）",
        "总体用量（含备用+留样，剂）",
    ]

    # 项目一：009（CpG+铝佐剂）
    row_009 = [
        "009\n（CpG+铝佐剂）",
        "1800例\n18-59岁 0,1月组 640\n18-59岁 0,2月组 640\n≥60岁组 520",
        "18-59岁试验组：0,1月 或 0,2月（2剂）\n≥60岁试验组：0,1,6月（3剂）",
        "10%\n（约412剂）",
        "20支\n（假设值）",
        4120,   # 1280 + 1280 + 1560
        412,    # 4120 × 10%
        4552,   # 4120 + 412 + 20
    ]

    # 项目二：009B（铝佐剂）
    row_009b = [
        "009B\n（铝佐剂）",
        "530例\n预试验 30 + 试验组 500",
        "0、1、6月（3剂）",
        "50支\n（约3.1%）",
        "20支\n（假设值）",
        1590,   # 90 + 1500
        50,     # 备用试验疫苗 50 支
        1660,   # 1590 + 50 + 20
    ]

    # ---------------------------------------------------------- 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="009系列Ⅲ期乙肝疫苗试验疫苗用量汇总")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ---------------------------------------------------------- 表头
    header_row = 2
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=text)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = WRAP_CENTER
    ws.row_dimensions[header_row].height = 34

    # ---------------------------------------------------------- 数据行
    for r, row in enumerate((row_009, row_009b), start=header_row + 1):
        for col, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.border = BORDER
            if col == 1:
                c.font = PROJ_FONT
                c.alignment = WRAP_CENTER
            elif col in (6, 7, 8):
                c.font = BODY_FONT
                c.alignment = CENTER
                c.number_format = "#,##0"
            else:
                c.font = BODY_FONT
                c.alignment = WRAP
        ws.row_dimensions[r].height = 76

    # ---------------------------------------------------------- 列宽
    widths = [14, 22, 30, 15, 12, 13, 13, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------------------------------------------------- 备注
    notes = [
        "备注：",
        "1. 统计口径：仅统计试验疫苗用量（对照疫苗除外）。",
        "2. 009 免疫程序：18-59岁试验组全程 2 剂（0,1月 或 0,2月）；≥60岁试验组全程 3 剂（0,1,6月）。全项目 2960 例（含对照组 1080 例），此处仅列试验组 1800 例。",
        "3. 009B 免疫程序：0、1、6月全程 3 剂。全项目主研究 1000 例（试验组 500 + 对照组 500），另有第一阶段预试验 30 例（仅接种试验疫苗 3 剂），已计入试验疫苗用量。",
        "4. 备用疫苗：009 为“约 10% 疫苗使用量”的备用疫苗（按试验疫苗基础用量 10% 计，约 412 剂）；009B 为固定 100 支（试验 50 + 对照 50），此处仅取试验疫苗备用 50 支。",
        "5. 留样用量为假设值：依据《药品生产质量管理规范》临床试验用药品附录（2022）第三十六条——“每批临床试验用药品均须留样，留样数量一般至少能完成两次全检，并至少保留一件最小包装成品”。假设重组乙肝疫苗一次全检量约 9~10 支、按 1 个生产批次计，取整 20 支/项目；实际应以质量标准全检量与生产批次数为准。",
        "6. 单位说明：本疫苗为单剂量包装（0.5 ml/支 = 1 次人用剂量），故 1 支 = 1 剂。",
        "7. 计算式：总体用量 = 基础用量 + 备用用量 + 留样用量。",
    ]
    note_start = header_row + 3  # 留一行空行
    ws.merge_cells(
        start_row=note_start, start_column=1,
        end_row=note_start, end_column=len(headers),
    )
    c = ws.cell(row=note_start, column=1, value=notes[0])
    c.font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
    for i, text in enumerate(notes[1:], start=1):
        r = note_start + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        c = ws.cell(row=r, column=1, value=text)
        c.font = NOTE_FONT
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 30 if len(text) > 40 else 18

    ws.sheet_view.showGridLines = False
    return wb


def main() -> int:
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    print(f"OK: 已生成 {OUTPUT_PATH}")
    print("  009  基础 4120 + 备用 412 + 留样 20 = 总体 4552 剂")
    print("  009B 基础 1590 + 备用 50  + 留样 20 = 总体 1660 剂")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
