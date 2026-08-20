import os

import openpyxl

os.chdir("review_materials")
wb = openpyxl.load_workbook("清洁残留计算所需信息.xlsx")
ws = wb.active

# Columns:
# 009: col C (idx 3 in openpyxl, but 1-indexed) -> Column 3
# 009B: col D -> Column 4
# 006: col E -> Column 5

# Row 7: 每日最大使用日剂量(LDD)
# Row 8: 产品活性成分最低日治疗剂量(MDD)

# 009
ws.cell(row=7, column=3).value = "1.0 mL"
ws.cell(row=8, column=3).value = "20 μg"

# 009B
ws.cell(row=7, column=4).value = "0.5 mL"
ws.cell(row=8, column=4).value = "20 μg"

# 006
ws.cell(row=7, column=5).value = "0.5 mL"
ws.cell(row=8, column=5).value = "50 μg"

wb.save("清洁残留计算所需信息.xlsx")
print("Excel updated successfully.")
